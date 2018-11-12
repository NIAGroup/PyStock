#####################################################################################
#### remember to include the "s" in the url so the proxy works effectively.      ####
#### REMEMBER TO INSTALL the alpha_vantage library for this code.                ####
#### alpha_vantage is a library we've found that's free and collects stock data. ####
#### other libraries needing installs : openpyxl & matplotlib                    ####
#### This code was originally written with Python 3.x, but it should work with   ####
#### python 2.x as well.                                                         ####
#####################################################################################

import json, requests, alpha_vantage, openpyxl, argparse, io
import matplotlib.pyplot as plt
import matplotlib
from matplotlib import dates
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

#### ~~ the method below parses the arguments provided by the user. In the event that ~~ ####
#### ~~ there are missing arguments or no arguments provided at all, the script will  ~~ ####
#### ~~ provide default arguments.                                                    ~~ ####
def getArgs():
    parser = argparse.ArgumentParser(description='stuff')
    parser.add_argument('-s')
    #args = parser.parse_args()
    parser.add_argument('-c')
    args = parser.parse_args()
    if args.s == None and args.c == None:
    	s = "INTC:NVDA"
    elif args.s == None and args.c != None:
    	s = "INTC:"+args.c
    elif args.s != None and args.c == None:
    	s = args.s+":INTC"
    elif args.s != None and args.c != None:
    	s = args.s+":"+args.c
    return s

#### ~~ The method below reverses the arrays so that the dates are incrementing.     ~~ ####
def getLeTicks(array_corrected_x):
	leTicks = []
	for t in range(0,len(array_corrected_x)):
		if t%4 == 0:
			leTicks.append(array_corrected_x[t])

	return leTicks

stock_symbols = getArgs().split(":")
stock_symbol_A = stock_symbols[0]
stock_symbol_B = stock_symbols[1]

print("stock A: "+stock_symbol_A)
print("stock B: "+stock_symbol_B)


#### ~~ The lines below are being used to prepare an excel file to store the found   ~~ ####
#### ~~ stock history.                                                               ~~ ####
row = 1
col = 1
wb = Workbook()
ws = wb.create_sheet("Main")
col_width = 25
ws.column_dimensions['A'].width = col_width
ws.column_dimensions['B'].width = col_width
ws.column_dimensions['C'].width = col_width
ws.column_dimensions['D'].width = col_width
ws.cell(row,col,"Company A - Date").font = Font(bold=True, size=14)
ws.cell(row,col+1,"Company A - Share Value").font = Font(bold=True, size=14)
ws.cell(row,col+2,"Company B - Date").font = Font(bold=True, size=14)
ws.cell(row,col+3,"Company B - Share Value").font = Font(bold=True, size=14)
ws.cell(row,col).alignment = Alignment(vertical="center", horizontal="center", wrap_text="True")
ws.cell(row,col+1).alignment = Alignment(vertical="center", horizontal="center", wrap_text="True")
ws.cell(row,col+2).alignment = Alignment(vertical="center", horizontal="center", wrap_text="True")
ws.cell(row,col+3).alignment = Alignment(vertical="center", horizontal="center", wrap_text="True")

share_array_A = []
year_array_A = []
array_corrected_y_A = []
array_corrected_x_A = []
share_array_B = []
year_array_B = []
array_corrected_y_B = []
array_corrected_x_B = []
API_URL = "https://www.alphavantage.co/query"

data_A = {
	"function": "SMA",
	"time_period": 10,
	"series_type": "open",
	"symbol": stock_symbol_A,
	"interval": "monthly",
	"apikey": "ksjendcksnfk",
}

data_B = {
	"function": "SMA",
	"time_period": 10,
	"series_type": "open",
	"symbol": stock_symbol_B,
	"interval": "monthly",
	"apikey": "ksjendcksnfk",
}

proxies = {
	"https" : "http://proxy-us.intel.com:911"
}
try:
    response_A = requests.get(API_URL, params=data_A,proxies=proxies)
    response_B = requests.get(API_URL, params=data_B,proxies=proxies)
except Exception:
    response_A = requests.get(API_URL, params=data_A)
    response_B = requests.get(API_URL, params=data_B)
arr_A = response_A.text.split("{")
arr_B = response_B.text.split("{")

js_A = json.loads(response_A.text)
js_B = json.loads(response_B.text)
d_A = js_A["Technical Analysis: SMA"]
d_B = js_B["Technical Analysis: SMA"]
row = 2
col = 1

#### ~~ The for loops below collect the target plot data from the json objects and add them ~~ ####
#### ~~ to arrays for the plot figure that will be created.                                 ~~ ####
for s in d_A:
	if(row>1):
	    stringer_a = str(s)
	    stringer_b = round(float(str(d_A[s]["SMA"])),2)
	    year_array_A.append(str(stringer_a))
	    share_array_A.append(float(stringer_b))
	    #print(stringer_a+"\t"+str(stringer_b))
	    ws.cell(row,col,stringer_a).alignment = Alignment(vertical="top", horizontal="center", wrap_text="True")
	    ws.cell(row,col+1,float(stringer_b)).alignment = Alignment(vertical="top", horizontal="center", wrap_text="True")
	    ws.cell(row,col+1).number_format = '$##0.00'
	row+=1
row = 2
col = 3

for s in d_B:
	if(row>1):
	    stringer_a = str(s)
	    stringer_b = round(float(d_B[s]["SMA"]),2)
	    year_array_B.append(str(stringer_a))
	    share_array_B.append(float(stringer_b))
	    #print(stringer_a+"\t"+str(stringer_b))
	    ws.cell(row,col,stringer_a).alignment = Alignment(vertical="top", horizontal="center", wrap_text="True")
	    ws.cell(row,col+1,float(stringer_b)).alignment = Alignment(vertical="top", horizontal="center", wrap_text="True")
	    ws.cell(row,col+1).number_format = '$##0.00'
	row+=1


wb.save("results.xlsx")
year_array_A =  matplotlib.dates.datestr2num(year_array_A)
year_array_B =  matplotlib.dates.datestr2num(year_array_B)
formatter = dates.DateFormatter('%Y-%m-%d')

#### ~~ The for loops below reverse the found data for the shares, as they were plotting in the ~~ ####
#### ~~ reverse (i.e. newest year to the lowest year).                                          ~~ ####
for x in range(len(year_array_A)-1,-1,-1):
	array_corrected_x_A.append(year_array_A[x])
for x in range(len(share_array_A)-1,-1,-1):
	array_corrected_y_A.append(share_array_A[x])

for x in range(len(year_array_B)-1,-1,-1):
	array_corrected_x_B.append(year_array_B[x])
for x in range(len(share_array_B)-1,-1,-1):
	array_corrected_y_B.append(share_array_B[x])

array_corrected_x_A = getLeTicks(array_corrected_x_A)
array_corrected_y_A = getLeTicks(array_corrected_y_A)
array_corrected_x_B = getLeTicks(array_corrected_x_B)
array_corrected_y_B = getLeTicks(array_corrected_y_B)

#### ~~ The lines below adjust some color formatting styles as well as labels text formatting. ~~ ####
fig = plt.figure() 
fig.patch.set_facecolor('xkcd:mint green')
plt.plot(array_corrected_x_A,array_corrected_y_A, color='darkorange',marker='o')
plt.plot(array_corrected_x_B,array_corrected_y_B, color='blue',marker='o')
plt.xlabel("Dates")
plt.ylabel("Share Value")
plt.title("Share Value Trends for: "+stock_symbol_A+" vs "+stock_symbol_B)
plt.grid()
the_ticks_A = array_corrected_x_A
the_ticks_B = array_corrected_x_B

#### ~~ This condition block was added to provide the full length of the x-axis. In some cases ~~ ####
#### ~~ the history of the primary share would be shorter, so the secondary would appear to be ~~ ####
#### ~~ plotting before the first data point (for date). So this was adjusted for formatting.  ~~ ####
if len(the_ticks_A) > len(the_ticks_B):
	plt.xticks(the_ticks_A)
else:
	plt.xticks(the_ticks_B)

plt.xticks(rotation=70)
ax = plt.gcf().axes[0] 

plt.legend((stock_symbol_A,stock_symbol_B))
ax.xaxis.set_major_formatter(formatter)
fmt = '${x:,.2f}'
tick = matplotlib.ticker.StrMethodFormatter(fmt)
ax.yaxis.set_major_formatter(tick)

#### ~~ The line below adjusts the dimensions of the plot to ensure all content is visible in ~~ ####
#### ~~ the figure window.                                                                    ~~ ####
plt.tight_layout()

#### ~~ The line below adjusts the width as the x axis labels appeared 'jumbled' ~~ #### 
fig.set_size_inches(round(fig.get_figwidth()*2.5,2), fig.get_figheight()*1.5, forward=True)

#### ~~ Saving as image. For some reason, placing this after adjusting the left parameter ~~ ####
#### ~~ leaves the figure blank. Also observed that the background is white upon saving   ~~ ####
#### ~~ but the plot lines still maintain their colors.                                   ~~ ####
plt.savefig('plot.png')

plt.subplots_adjust(left=0.07)
plt.show()

plt.close()