import datetime
from django.db import models
from django.utils import timezone
import json, requests, alpha_vantage, openpyxl, argparse, io
import matplotlib.pyplot as plt
import matplotlib
from matplotlib import dates
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class test:
    x = 1
    def printer(self):
        self.x+=1
        print(self.x)
        return self.x
#### ~~ The method below reverses the arrays so that the dates are incrementing.     ~~ ####
def getLeTicks(array_corrected_x):
    leTicks = []
    for t in range(0,len(array_corrected_x)):
        if t%4 == 0:
            leTicks.append(array_corrected_x[t])

    return leTicks

class plot_comparison(models.Model):
    def showPlot(self):
        print("working")
        formatter = dates.DateFormatter('%Y-%m-%d')
        array_corrected_x_A = self.array_corrected_x_A
        array_corrected_x_B = self.array_corrected_x_B
        array_corrected_y_A = self.array_corrected_y_A
        array_corrected_y_B = self.array_corrected_y_B
        #### ~~ The lines below adjust some color formatting styles as well as labels text formatting. ~~ ####
        fig = plt.figure()
        fig.patch.set_facecolor('xkcd:mint green')
        plt.plot(array_corrected_x_A, array_corrected_y_A, color='darkorange', marker='o')
        plt.plot(array_corrected_x_B, array_corrected_y_B, color='blue', marker='o')
        plt.xlabel("Dates")
        plt.ylabel("Share Value")
        plt.title("Share Value Trends for: " + self.stock_symbol_A + " vs " + self.stock_symbol_B)
        plt.grid()
        the_ticks_A = array_corrected_x_A
        the_ticks_B = array_corrected_x_B

        #### ~~ This condition block was added to provide the full length of the x-axis. In some cases ~~ ####
        #### ~~ the history of the primary share would be shorter, so the secondary would appear to be ~~ ####
        #### ~~ plotting before the first data point (for date). So this was adjusted for formatting.  ~~ ####
        if len(the_ticks_A) > len(the_ticks_B):
            plt.xticks(the_ticks_A)
        else:
            plt.xticks(the_ticks_B)

        plt.xticks(rotation=70)
        ax = plt.gcf().axes[0]

        plt.legend((self.stock_symbol_A, self.stock_symbol_B))
        ax.xaxis.set_major_formatter(formatter)
        fmt = '${x:,.2f}'
        tick = matplotlib.ticker.StrMethodFormatter(fmt)
        ax.yaxis.set_major_formatter(tick)

        #### ~~ The line below adjusts the dimensions of the plot to ensure all content is visible in ~~ ####
        #### ~~ the figure window.                                                                    ~~ ####
        plt.tight_layout()

        #### ~~ The line below adjusts the width as the x axis labels appeared 'jumbled' ~~ ####
        fig.set_size_inches(round(fig.get_figwidth() * 2.5, 2), fig.get_figheight() * 1.5, forward=True)

        #### ~~ Saving as image. For some reason, placing this after adjusting the left parameter ~~ ####
        #### ~~ leaves the figure blank. Also observed that the background is white upon saving   ~~ ####
        #### ~~ but the plot lines still maintain their colors.                                   ~~ ####
        plt.savefig('plot.png')

        plt.subplots_adjust(left=0.07)
        plt.show()

    def __init__(self, stock_symbol_A, stock_symbol_B):
        #### ~~ the method below parses the arguments provided by the user. In the event that ~~ ####
        #### ~~ there are missing arguments or no arguments provided at all, the script will  ~~ ####
        #### ~~ provide default arguments.                                                    ~~ ####
        # stock_symbols = getArgs().split(":")
        # stock_symbol_A = stock_symbols[0]
        # stock_symbol_B = stock_symbols[1]
        self.stock_symbol_A = stock_symbol_A
        self.stock_symbol_B = stock_symbol_B
        print("stock A: " + stock_symbol_A)
        print("stock B: " + stock_symbol_B)

        #### ~~ The lines below are being used to prepare an excel file to store the found   ~~ ####
        #### ~~ stock history.                                                               ~~ ####
        row = 1
        col = 1
        wb = Workbook()
        ws = wb.create_sheet("Main")
        col_width = 25
        ws.column_dimensions['A'].width = col_width
        ws.column_dimensions['B'].width = col_width
        ws.column_dimensions['C'].width = col_width
        ws.column_dimensions['D'].width = col_width
        ws.cell(row, col, "Company A - Date").font = Font(bold=True, size=14)
        ws.cell(row, col + 1, "Company A - Share Value").font = Font(bold=True, size=14)
        ws.cell(row, col + 2, "Company B - Date").font = Font(bold=True, size=14)
        ws.cell(row, col + 3, "Company B - Share Value").font = Font(bold=True, size=14)
        ws.cell(row, col).alignment = Alignment(vertical="center", horizontal="center", wrap_text="True")
        ws.cell(row, col + 1).alignment = Alignment(vertical="center", horizontal="center", wrap_text="True")
        ws.cell(row, col + 2).alignment = Alignment(vertical="center", horizontal="center", wrap_text="True")
        ws.cell(row, col + 3).alignment = Alignment(vertical="center", horizontal="center", wrap_text="True")

        share_array_A = []
        year_array_A = []
        array_corrected_y_A = []
        array_corrected_x_A = []
        share_array_B = []
        year_array_B = []
        array_corrected_y_B = []
        array_corrected_x_B = []
        API_URL = "https://www.alphavantage.co/query"

        data_A = {
            "function": "SMA",
            "time_period": 10,
            "series_type": "open",
            "symbol": stock_symbol_A,
            "interval": "monthly",
            "apikey": "ksjendcksnfk",
        }

        data_B = {
            "function": "SMA",
            "time_period": 10,
            "series_type": "open",
            "symbol": stock_symbol_B,
            "interval": "monthly",
            "apikey": "ksjendcksnfk",
        }

        proxies = {
            "https": "http://proxy-us.intel.com:911"
        }
        try:
            response_A = requests.get(API_URL, params=data_A, proxies=proxies)
            response_B = requests.get(API_URL, params=data_B, proxies=proxies)
        except Exception:
            response_A = requests.get(API_URL, params=data_A)
            response_B = requests.get(API_URL, params=data_B)
        arr_A = response_A.text.split("{")
        arr_B = response_B.text.split("{")

        js_A = json.loads(response_A.text)
        js_B = json.loads(response_B.text)
        d_A = js_A["Technical Analysis: SMA"]
        d_B = js_B["Technical Analysis: SMA"]
        row = 2
        col = 1

        #### ~~ The for loops below collect the target plot data from the json objects and add them ~~ ####
        #### ~~ to arrays for the plot figure that will be created.                                 ~~ ####
        for s in d_A:
            if (row > 1):
                stringer_a = str(s)
                stringer_b = round(float(str(d_A[s]["SMA"])), 2)
                year_array_A.append(str(stringer_a))
                share_array_A.append(float(stringer_b))
                # print(stringer_a+"\t"+str(stringer_b))
                ws.cell(row, col, stringer_a).alignment = Alignment(vertical="top", horizontal="center",
                                                                    wrap_text="True")
                ws.cell(row, col + 1, float(stringer_b)).alignment = Alignment(vertical="top", horizontal="center",
                                                                               wrap_text="True")
                ws.cell(row, col + 1).number_format = '$##0.00'
            row += 1
        row = 2
        col = 3

        for s in d_B:
            if (row > 1):
                stringer_a = str(s)
                stringer_b = round(float(d_B[s]["SMA"]), 2)
                year_array_B.append(str(stringer_a))
                share_array_B.append(float(stringer_b))
                # print(stringer_a+"\t"+str(stringer_b))
                ws.cell(row, col, stringer_a).alignment = Alignment(vertical="top", horizontal="center",
                                                                    wrap_text="True")
                ws.cell(row, col + 1, float(stringer_b)).alignment = Alignment(vertical="top", horizontal="center",
                                                                               wrap_text="True")
                ws.cell(row, col + 1).number_format = '$##0.00'
            row += 1

        wb.save("results.xlsx")
        year_array_A = matplotlib.dates.datestr2num(year_array_A)
        year_array_B = matplotlib.dates.datestr2num(year_array_B)
        formatter = dates.DateFormatter('%Y-%m-%d')

        #### ~~ The for loops below reverse the found data for the shares, as they were plotting in the ~~ ####
        #### ~~ reverse (i.e. newest year to the lowest year).                                          ~~ ####
        for x in range(len(year_array_A) - 1, -1, -1):
            array_corrected_x_A.append(year_array_A[x])
        for x in range(len(share_array_A) - 1, -1, -1):
            array_corrected_y_A.append(share_array_A[x])

        for x in range(len(year_array_B) - 1, -1, -1):
            array_corrected_x_B.append(year_array_B[x])
        for x in range(len(share_array_B) - 1, -1, -1):
            array_corrected_y_B.append(share_array_B[x])

        array_corrected_x_A = getLeTicks(array_corrected_x_A)
        array_corrected_y_A = getLeTicks(array_corrected_y_A)
        array_corrected_x_B = getLeTicks(array_corrected_x_B)
        array_corrected_y_B = getLeTicks(array_corrected_y_B)
        self.array_corrected_x_A = array_corrected_x_A
        self.array_corrected_x_B = array_corrected_x_B
        self.array_corrected_y_A = array_corrected_y_A
        self.array_corrected_y_B = array_corrected_y_B

    def __str__(self):
        return self.array_corrected_x_A

class Question(models.Model):
    question_text = models.CharField(max_length=200)
    pub_date = models.DateTimeField('date published')
    #plot = plot_comparison("AAPL","GOOGL")
    x = test()
    def __str__(self):
        return self.question_text 

    def was_published_recently(self):
        now = timezone.now()
        return now - datetime.timedelta(days=1) <= self.pub_date <= now
    class test:

        def printer(x):
            x+=1
            print(x)
            return x
        x = 1
        printer(x)



class Choice(models.Model):
    question = models.ForeignKey(Question, on_delete=models.CASCADE)
    choice_text = models.CharField(max_length=200)
    votes = models.IntegerField(default=0)

    def __str__(self):
        return self.choice_text
        
class Plot(models.Model):
    image_placeholder = models.ImageField(upload_to=None,height_field=200,width_field=200,max_length=100)
    
    def __str__(self):
        return "we got an image"

# Create your models here.
