import sys, json, os, openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

xlsx_file_name = "results.xlsx" 
json_file_name = "results.json"

wb = load_workbook(xlsx_file_name)
row = 2
col = 1
Shares_dict = {}
Shares_dict["Stock Values"] = []
ws1 = wb['Main']

def jsonCreation(dict_keys, wb, ws1, row, col, json_file_name, Shares_dict):
	upper_limit = row 
	while row > 1 :
		Shares_dict["Stock Values"].append({
		"Value_"+str(upper_limit-row) : "Value_"+str(upper_limit-row),
		})
		Shares_dict["Stock Values"][upper_limit-row]["Value_"+str(upper_limit-row)] = []
		for c in range(0,len(dict_keys)):
			if ws1.cell(row,c+1).value == "" or ws1.cell(row,c+1).value == None:
				Shares_dict["Stock Values"][upper_limit-row]["Value_"+str(upper_limit-row)].append({dict_keys[c] : ""})
				#print(ws1.cell(row,c+1).value)
			else:
				Shares_dict["Stock Values"][upper_limit-row]["Value_"+str(upper_limit-row)].append({dict_keys[c] : str(ws1.cell(row,c+1).value)})
		row-=1

def getKeys(wb, ws1):
	row = 1
	col = 1
	dict_keys = []
	while(ws1.cell(row,col).value != None and ws1.cell(row,col).value != ""):
		dict_keys.append(ws1.cell(row,col).value)
		col+=1
	col = 1
	while(ws1.cell(row,col).value != None and ws1.cell(row,col).value != ""):
		row+=1
	dict_keys.append(row - 1)
	return dict_keys

dict_keys = getKeys(wb, ws1)
row = dict_keys[len(dict_keys)-1]
del dict_keys[len(dict_keys)-1]
col = 1
jsonCreation(dict_keys, wb, ws1, row, col, json_file_name, Shares_dict)

#print(json.dumps(Shares_dict,indent=4))
library = json.dumps(Shares_dict,indent=4)
file = open(json_file_name, 'w')
file.write(library)
file.close()