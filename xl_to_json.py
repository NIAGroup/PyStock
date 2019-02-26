import sys, json, os, openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

xlsx_file_name = "results.xlsx" 
json_file_name = "results.json"

wb = load_workbook(xlsx_file_name)
row = 2
col = 1
Dispensaries = {}
Dispensaries["Dispensaries"] = []

ws1 = wb['Dispensary ']
while ws1.cell(row,col).value != "" and ws1.cell(row,col).value != None:
	#print(ws1.cell(row,col).value)
	Dispensaries["Dispensaries"].append({
	"Name" : str(ws1.cell(row,col).value),
	"Address" : ws1.cell(row,2).value,
	"Phone" : ws1.cell(row,3).value,
	"Hours of Operation" : ws1.cell(row,4).value,
	"City and State" : ws1.cell(row,5).value,
	"Type of Business" : ws1.cell(row,6).value
	})
	row+=1
#print(json.dumps(Dispensaries,indent=4))
library = json.dumps(Dispensaries,indent=4)
file = open(json_file_name, 'w')
file.write(library)
file.close()